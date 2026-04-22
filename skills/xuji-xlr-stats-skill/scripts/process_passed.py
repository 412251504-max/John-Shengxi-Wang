"""
薛记学习率统计处理脚本
输入: 项目地图情况表.xlsx, 员工账号管理.xlsx, 月架构.xlsx
输出: 单项目完成情况表.xlsx, 区域学习率统计.xlsx, 东区学习率统计.xlsx, 催学名单.xlsx
"""

import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys, os

# ── 小区负责人 → 成功部 映射 ──
XQ_TO_BU = {
    '周亭': '合伙人成功一部',
    '赵丹': '合伙人成功二部',
    '岳辉': '合伙人成功三部',
    '王刚': '合伙人成功四部',
}
XQ_ORDER = ['周亭', '赵丹', '岳辉', '王刚']

BU_ORDER = ['合伙人成功一部', '合伙人成功二部', '合伙人成功三部', '合伙人成功四部']
BU_COMPANY_ORDER = {
    '合伙人成功一部': ['安徽销售中心', '湖北销售中心', '江苏销售中心', '山西销售中心', '陕西销售中心', '重庆销售中心'],
    '合伙人成功二部': ['江西销售中心', '浙江销售中心'],
    '合伙人成功三部': ['江苏销售中心', '上海销售中心'],
    '合伙人成功四部': ['江苏销售中心'],
}

# 催学名单配色（每位负责人一个主色）
XQ_COLORS = {'周亭': '3A7D44', '赵丹': '1B7A8C', '岳辉': 'B8862A', '王刚': '6B3FA0'}
XQ_ALT    = {'周亭': 'DDF0E0', '赵丹': 'D4EEF2', '岳辉': 'F5E9CC', '王刚': 'E8D8F7'}

# ── 样式工具 ──
THIN = Border(left=Side(style='thin'), right=Side(style='thin'),
              top=Side(style='thin'), bottom=Side(style='thin'))

def sc(ws, row, col, value, bold=False, color=None, bg=None, fmt=None,
        align='center', wrap=True, size=10):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name='Arial', bold=bold, size=size,
                     color=color if color else '000000')
    if bg:
        cell.fill = PatternFill('solid', start_color=bg)
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    cell.border = THIN
    if fmt:
        cell.number_format = fmt
    return cell


# ── 数据处理 ──

def load_data(map_file, exclude_file, arch_file):
    df_map     = pd.read_excel(map_file, header=1)
    df_exclude = pd.read_excel(exclude_file, header=1)
    df_arch    = pd.read_excel(arch_file, sheet_name='Sheet', header=0)
    return df_map, df_exclude, df_arch


def filter_data(df_map, df_exclude):
    exclude_ids = set(df_exclude['员工编号'].astype(str).str.strip())
    df_map['_eid'] = df_map['员工编号'].astype(str).str.strip()
    is_store = df_map['部门名称'].astype(str).str.match(r'^[A-Z]\d{3}')
    df = df_map[is_store & ~df_map['_eid'].isin(exclude_ids)].copy()
    df['_店编码'] = df['部门名称'].astype(str).str.extract(r'^([A-Z]\d+)')[0]
    df['_门店名'] = df['部门名称'].astype(str).str.replace(r'^[A-Z]\d+-', '', regex=True)
    df['_完成']  = df['通过状态'] == '已通过'  # 口径：已通过

    # 白名单：只保留一线岗位
    WHITELIST = {'营业员', '店经理', '店经理助理', '储备店经理', '训练员', '训练组长', '销售专员', '管理培训生', '实习生'}
    df = df[df['岗位名称'].isin(WHITELIST)]
    return df


def enrich_with_arch(df, df_arch):
    direct = df_arch.set_index('机构编码')['小区负责人'].to_dict()
    fallback = {(r['分公司'], r['区域']): r['小区负责人'] for _, r in df_arch.iterrows()}
    def get_xq(row):
        v = direct.get(row['_店编码'], '')
        return v if v else fallback.get((row['部门名称3'], row['部门名称4']), '')
    df['小区负责人'] = df.apply(get_xq, axis=1)

    # 手动覆盖：架构表未收录的区域
    MANUAL_REGION_MAP = {
        ('陕西销售中心', '延安区'): '周亭',
        ('湖北销售中心', '孝感区'): '周亭',
        ('江西销售中心', '吉安区'): '赵丹',
    }
    for (company, region), xq in MANUAL_REGION_MAP.items():
        mask = (df['小区负责人'] == '') & (df['部门名称3'] == company) & (df['部门名称4'] == region)
        df.loc[mask, '小区负责人'] = xq

    return df


# ── 报表构建 ──

def build_single_project_table(df):
    out = df[['项目名称','员工编号','员工姓名','入职日期','部门名称2','部门名称3','部门名称4']].copy()
    out['部门名称']     = df['_店编码']
    out['小区负责人']   = df['小区负责人']
    out['门店']         = df['_门店名']
    out['岗位名称']     = df['岗位名称']
    out['职位名称']     = df['职位名称']
    out['当前必修任务'] = df['当前必修任务']
    out['必修任务数']   = df['必修任务数']
    out['已完成必修任务数'] = df['已完成必修任务数']
    out['项目完成进度'] = df['项目完成进度']
    out['完成状态']     = df['完成状态']
    out['项目通过进度'] = df['项目通过进度']
    out['通过状态']     = df['通过状态']
    out['就职状态']     = df['就职状态']
    return out.sort_values(['部门名称3','部门名称4','完成状态'])


def build_region_stats(df):
    g = df[df['部门名称3'] != '东区事业部'].groupby(['部门名称3','部门名称4']).agg(
        已完成人数=('_完成','sum'), 总人数=('员工编号','count')).reset_index()
    g['未完成人数'] = g['总人数'] - g['已完成人数']
    g['学习率']     = g['已完成人数'] / g['总人数']
    g.rename(columns={'部门名称3':'销售中心','部门名称4':'区域'}, inplace=True)
    total = pd.DataFrame([{
        '销售中心':'区域合计','区域':'区域合计',
        '未完成人数': g['未完成人数'].sum(),
        '已完成人数': g['已完成人数'].sum(),
        '总人数':     g['总人数'].sum(),
        '学习率':     g['已完成人数'].sum() / g['总人数'].sum()
    }])
    return pd.concat([g[['销售中心','区域','未完成人数','已完成人数','总人数','学习率']], total], ignore_index=True)


def build_east_stats(df):
    rows = []
    direct  = df[df['部门名称3'] == '东区事业部']
    partner = df[df['部门名称3'] != '东区事业部'].copy()
    partner['_成功部'] = partner['小区负责人'].map(XQ_TO_BU).fillna('未知')

    dg = direct.groupby('部门名称4').agg(已完成=('_完成','sum'), 总人数=('员工编号','count')).reset_index()
    for bu in BU_ORDER:
        r = dg[dg['部门名称4'] == bu]
        if len(r):
            rr = r.iloc[0]
            rows.append(('直营销售中心', bu, int(rr['总人数']-rr['已完成']), int(rr['已完成']), int(rr['总人数']), rr['已完成']/rr['总人数']))
    rows.append(('直营区统计', None, int(len(direct)-direct['_完成'].sum()), int(direct['_完成'].sum()), len(direct), direct['_完成'].mean()))

    pg = partner.groupby(['_成功部','部门名称3']).agg(已完成=('_完成','sum'), 总人数=('员工编号','count')).reset_index()
    for bu in BU_ORDER:
        seen, sub = set(), pg[pg['_成功部'] == bu]
        for company in BU_COMPANY_ORDER.get(bu, []):
            r = sub[sub['部门名称3'] == company]
            if len(r):
                rr = r.iloc[0]
                rows.append((bu, company, int(rr['总人数']-rr['已完成']), int(rr['已完成']), int(rr['总人数']), rr['已完成']/rr['总人数']))
                seen.add(company)
        for _, rr in sub[~sub['部门名称3'].isin(seen)].iterrows():
            rows.append((bu, rr['部门名称3'], int(rr['总人数']-rr['已完成']), int(rr['已完成']), int(rr['总人数']), rr['已完成']/rr['总人数']))

    rows.append(('合伙区统计', None, int(len(partner)-partner['_完成'].sum()), int(partner['_完成'].sum()), len(partner), partner['_完成'].mean()))
    return rows


def build_reminder_dict(df):
    """返回 {小区负责人: 未完成 DataFrame}"""
    df_inc = df[~df['_完成']].copy()
    status_ord = {'进行中': 0, '未开始': 1}
    df_inc['_sort'] = df_inc['完成状态'].map(status_ord).fillna(2)
    result = {}
    for xq in XQ_ORDER:
        sub = df_inc[df_inc['小区负责人'] == xq].sort_values(
            ['部门名称3','部门名称4','_门店名','_sort'])
        result[xq] = sub
    return result


# ── Excel 写出 ──

def _write_header_row(ws, row, headers, bg, cols_widths=None):
    for i, h in enumerate(headers, 1):
        sc(ws, row, i, h, bold=True, color='FFFFFF', bg=bg)
    if cols_widths:
        for i, w in enumerate(cols_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 18


def write_region_stats(df, project_name, output_path):
    wb = Workbook(); ws = wb.active; ws.title = 'Sheet1'
    ws.merge_cells('A1:F1')
    sc(ws, 1, 1, project_name, bold=True, color='FFFFFF', bg='1F4E79', size=11)
    _write_header_row(ws, 2, ['销售中心','区域','未完成人数','已完成人数','总人数','学习率'],
                      '4472C4', [16,12,12,12,10,10])
    ws.row_dimensions[1].height = 22

    data = df[df['销售中心'] != '区域合计']
    row_n = 3; prev = None; cs_row = 3
    for _, r in data.iterrows():
        company = r['销售中心']
        if company != prev:
            if prev and row_n - cs_row > 1:
                ws.merge_cells(f'A{cs_row}:A{row_n-1}')
                ws.cell(cs_row,1).alignment = Alignment(horizontal='center', vertical='center')
            cs_row = row_n; prev = company
        alt = 'DEEAF1' if row_n % 2 == 0 else None
        sc(ws, row_n, 1, company, bg=alt)
        sc(ws, row_n, 2, r['区域'], bg=alt)
        sc(ws, row_n, 3, int(r['未完成人数']), bg=alt)
        sc(ws, row_n, 4, int(r['已完成人数']), bg=alt)
        sc(ws, row_n, 5, int(r['总人数']), bg=alt)
        sc(ws, row_n, 6, r['学习率'], bg=alt, fmt='0%')
        row_n += 1
    if row_n - cs_row > 1:
        ws.merge_cells(f'A{cs_row}:A{row_n-1}')
        ws.cell(cs_row,1).alignment = Alignment(horizontal='center', vertical='center')
    total = df[df['销售中心'] == '区域合计'].iloc[0]
    for i, v in enumerate([('区域合计',None),('区域合计',None),(int(total['未完成人数']),None),
                            (int(total['已完成人数']),None),(int(total['总人数']),None),(total['学习率'],'0%')], 1):
        sc(ws, row_n, i, v[0], bold=True, bg='BDD7EE', fmt=v[1])
    wb.save(output_path)


def write_east_stats(rows, project_name, output_path):
    wb = Workbook(); ws = wb.active; ws.title = 'Sheet1'
    ws.merge_cells('A1:F1')
    sc(ws, 1, 1, project_name, bold=True, color='FFFFFF', bg='1F4E79', size=11)
    _write_header_row(ws, 2, ['部门','销售中心','未完成人数','已完成人数','总人数','学习率'],
                      '4472C4', [18,16,12,12,10,10])
    ws.row_dimensions[1].height = 22

    row_n = 3
    for r in rows:
        bu, company, not_done, done, total, rate = r
        is_sub = company is None
        bg = 'BDD7EE' if is_sub else ('DEEAF1' if row_n % 2 == 0 else None)
        bold = is_sub
        sc(ws, row_n, 1, bu,       bold=bold, bg=bg)
        sc(ws, row_n, 2, company,  bold=bold, bg=bg)
        sc(ws, row_n, 3, not_done, bold=bold, bg=bg)
        sc(ws, row_n, 4, done,     bold=bold, bg=bg)
        sc(ws, row_n, 5, total,    bold=bold, bg=bg)
        sc(ws, row_n, 6, rate,     bold=bold, bg=bg, fmt='0%')
        row_n += 1

    # 合并直营列
    dr = [i+3 for i, r in enumerate(rows) if r[0] == '直营销售中心']
    if len(dr) > 1:
        ws.merge_cells(f'A{dr[0]}:A{dr[-1]}')
        ws.cell(dr[0],1).alignment = Alignment(horizontal='center', vertical='center')
    for bu in BU_ORDER:
        br = [i+3 for i, r in enumerate(rows) if r[0] == bu and r[1] is not None]
        if len(br) > 1:
            ws.merge_cells(f'A{br[0]}:A{br[-1]}')
            ws.cell(br[0],1).alignment = Alignment(horizontal='center', vertical='center')
    wb.save(output_path)


def write_single_project(df, project_name, output_path):
    wb = Workbook(); ws = wb.active; ws.title = 'Sheet1'
    headers = ['项目名称','员工编号','员工姓名','入职日期','部门名称2','部门名称3',
               '部门名称4','部门名称','小区负责人','门店','岗位名称','职位名称',
               '当前必修任务','必修任务数','已完成必修任务数','项目完成进度',
               '完成状态','项目通过进度','通过状态','就职状态']
    ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
    sc(ws, 1, 1, project_name, bold=True, color='FFFFFF', bg='1F4E79', size=11)
    _write_header_row(ws, 2, headers, '4472C4',
                      [30,10,8,12,16,14,10,8,8,16,10,12,30,8,12,10,8,10,8])
    ws.row_dimensions[1].height = 22
    for seq, (_, r) in enumerate(df.iterrows(), 3):
        alt = 'DEEAF1' if seq % 2 == 0 else None
        for ci, h in enumerate(headers, 1):
            fmt = '0%' if h in ('项目完成进度','项目通过进度') else None
            sc(ws, seq, ci, r.get(h,''), bg=alt, fmt=fmt)
    ws.freeze_panes = 'A3'
    wb.save(output_path)



def build_failed_exam_list(df):
    """生成完成但考核未通过名单：完成状态=已完成 且 通过状态≠已通过"""
    mask = (df['完成状态'] == '已完成') & (df['通过状态'] != '已通过')
    df_fail = df[mask].copy()
    df_fail = df_fail.sort_values(['小区负责人', '部门名称3', '部门名称4', '_门店名'])
    return df_fail


def write_failed_exam_list(df_fail, project_name, output_path):
    """写出考核未通过名单，按小区负责人分Sheet"""
    wb = Workbook()
    wb.remove(wb.active)

    headers = ['序号', '小区负责人', '分公司', '区域', '门店', '员工编号', '员工姓名', '岗位名称', '完成状态', '通过状态']
    col_widths = [6, 8, 14, 10, 20, 10, 8, 10, 10, 10]

    # 汇总 Sheet
    ws_sum = wb.create_sheet('汇总')
    ws_sum.merge_cells(f'A1:{get_column_letter(len(headers))}1')
    sc(ws_sum, 1, 1, f'考核未通过名单  ·  {project_name}  ·  共 {len(df_fail)} 人',
       bold=True, color='FFFFFF', bg='A32D2D', size=11)
    ws_sum.row_dimensions[1].height = 24
    for i, h in enumerate(headers, 1):
        sc(ws_sum, 2, i, h, bold=True, color='FFFFFF', bg='C55A11')
    ws_sum.row_dimensions[2].height = 18

    for seq, (_, r) in enumerate(df_fail.iterrows(), 1):
        bg = 'FDEBD0' if seq % 2 == 0 else None
        try: prog_val = r['_门店名']
        except: prog_val = ''
        sc(ws_sum, seq+2, 1, seq, bg=bg)
        sc(ws_sum, seq+2, 2, r['小区负责人'], bg=bg)
        sc(ws_sum, seq+2, 3, r['部门名称3'], bg=bg)
        sc(ws_sum, seq+2, 4, r['部门名称4'], bg=bg)
        sc(ws_sum, seq+2, 5, r['_门店名'], bg=bg, align='left')
        sc(ws_sum, seq+2, 6, str(r['员工编号']), bg=bg)
        sc(ws_sum, seq+2, 7, r['员工姓名'], bg=bg)
        sc(ws_sum, seq+2, 8, r['岗位名称'], bg=bg)
        sc(ws_sum, seq+2, 9, r['完成状态'], bg='E2EFDA', color='375623')
        sc(ws_sum, seq+2, 10, r['通过状态'] if not str(r['通过状态']) == 'nan' else '未通过', bg='FADBD8', color='A32D2D')

    for i, w in enumerate(col_widths, 1):
        ws_sum.column_dimensions[get_column_letter(i)].width = w
    ws_sum.freeze_panes = 'A3'

    wb.save(output_path)


def write_reminder_list(reminder_dict, project_name, df_all, output_path):
    """
    催学名单：4个 Sheet（每位小区负责人），+ 1个汇总 Sheet
    df_all 用于计算汇总页的总人数/学习率
    """
    wb = Workbook(); wb.remove(wb.active)

    headers = ['序号','分公司','区域','门店','员工编号','员工姓名','岗位名称','完成状态','进度']
    col_widths = [6, 14, 10, 20, 10, 8, 10, 10, 8]

    # ── 汇总 Sheet ──
    ws_sum = wb.create_sheet('汇总')
    ws_sum.merge_cells('A1:G1')
    sc(ws_sum, 1, 1, f'催学名单汇总  ·  {project_name}',
       bold=True, color='FFFFFF', bg='1F4E79', size=11)
    ws_sum.row_dimensions[1].height = 26
    _write_header_row(ws_sum, 2,
                      ['小区负责人','成功部','未完成','已完成','总人数','学习率','备注'],
                      '2F5496', [14,16,10,10,10,10,20])

    # ── 各小区 Sheet ──
    for xq in XQ_ORDER:
        df_xq = reminder_dict[xq]
        color = XQ_COLORS[xq]
        alt   = XQ_ALT[xq]

        ws = wb.create_sheet(xq)
        # 标题行
        ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
        sc(ws, 1, 1,
           f'{xq} 小区  ·  {project_name}  ·  未完成 {len(df_xq)} 人',
           bold=True, color='FFFFFF', bg=color, size=11,
           align='left')
        ws.row_dimensions[1].height = 24
        _write_header_row(ws, 2, headers, color, col_widths)

        for seq, (_, r) in enumerate(df_xq.iterrows(), 1):
            row_n = seq + 2
            bg = alt if seq % 2 == 0 else None
            status = r['完成状态']
            # 进行中：橙色高亮
            status_bg   = 'FFE699' if status == '进行中' else bg
            status_color= '7F4F00' if status == '进行中' else '000000'
            try:
                prog = float(r['项目完成进度'])
            except (ValueError, TypeError):
                prog = 0.0

            sc(ws, row_n, 1, seq,             bg=bg)
            sc(ws, row_n, 2, r['部门名称3'],   bg=bg)
            sc(ws, row_n, 3, r['部门名称4'],   bg=bg)
            sc(ws, row_n, 4, r['_门店名'],     bg=bg, align='left')
            sc(ws, row_n, 5, str(r['员工编号']),bg=bg)
            sc(ws, row_n, 6, r['员工姓名'],    bg=bg)
            sc(ws, row_n, 7, r['岗位名称'],    bg=bg)
            sc(ws, row_n, 8, status,           bg=status_bg, color=status_color)
            sc(ws, row_n, 9, prog,             bg=bg, fmt='0%')

        ws.freeze_panes = 'A3'
        ws.auto_filter.ref = f'A2:{get_column_letter(len(headers))}2'

        # 写汇总行
        all_xq = df_all[df_all['小区负责人'] == xq]
        done   = int(all_xq['_完成'].sum())
        total  = len(all_xq)
        not_done = total - done
        rate   = done / total if total else 0
        row_i  = XQ_ORDER.index(xq) + 3
        xq_alt = XQ_ALT[xq]
        sc(ws_sum, row_i, 1, xq,          bg=xq_alt, bold=True)
        sc(ws_sum, row_i, 2, XQ_TO_BU[xq],bg=xq_alt)
        sc(ws_sum, row_i, 3, not_done,    bg=xq_alt)
        sc(ws_sum, row_i, 4, done,        bg=xq_alt)
        sc(ws_sum, row_i, 5, total,       bg=xq_alt)
        sc(ws_sum, row_i, 6, rate,        bg=xq_alt, fmt='0.0%')
        sc(ws_sum, row_i, 7, f'名单见「{xq}」Sheet', bg=xq_alt, align='left')

    # 汇总合计行
    total_row = len(XQ_ORDER) + 3
    not_done_total = sum(len(reminder_dict[xq]) for xq in XQ_ORDER)
    done_total = int(df_all['_完成'].sum())
    grand_total = len(df_all)
    sc(ws_sum, total_row, 1, '合计', bold=True, bg='BDD7EE')
    sc(ws_sum, total_row, 2, '',     bold=True, bg='BDD7EE')
    sc(ws_sum, total_row, 3, not_done_total, bold=True, bg='BDD7EE')
    sc(ws_sum, total_row, 4, done_total,     bold=True, bg='BDD7EE')
    sc(ws_sum, total_row, 5, grand_total,    bold=True, bg='BDD7EE')
    sc(ws_sum, total_row, 6, done_total/grand_total if grand_total else 0,
       bold=True, bg='BDD7EE', fmt='0.0%')
    sc(ws_sum, total_row, 7, '', bg='BDD7EE')

    wb.save(output_path)


# ── 主入口 ──

def run(map_file, exclude_file, arch_file, output_dir='.'):
    print("正在读取数据...")
    df_map, df_exclude, df_arch = load_data(map_file, exclude_file, arch_file)

    print("正在过滤数据...")
    df = filter_data(df_map, df_exclude)
    df = enrich_with_arch(df, df_arch)

    project_name = df['项目名称'].iloc[0] if len(df) else '未知项目'
    print(f"项目: {project_name}")
    print(f"有效人员: {len(df)} 人  |  已完成: {df['_完成'].sum()} 人  |  学习率: {df['_完成'].mean():.1%}")

    os.makedirs(output_dir, exist_ok=True)

    print("→ 生成单项目完成情况表...")
    single_path = os.path.join(output_dir, f"{project_name}.xlsx")
    write_single_project(build_single_project_table(df), project_name, single_path)

    print("→ 生成区域学习率统计...")
    region_path = os.path.join(output_dir, "区域学习率统计.xlsx")
    write_region_stats(build_region_stats(df), project_name, region_path)

    print("→ 生成东区学习率统计...")
    east_path = os.path.join(output_dir, "东区学习率统计.xlsx")
    write_east_stats(build_east_stats(df), project_name, east_path)

    print("→ 生成催学名单...")
    reminder_path = os.path.join(output_dir, "催学名单.xlsx")
    write_reminder_list(build_reminder_dict(df), project_name, df, reminder_path)

    print("→ 生成考核未通过名单...")
    df_fail = build_failed_exam_list(df)
    failed_path = os.path.join(output_dir, "考核未通过名单.xlsx")
    if len(df_fail) > 0:
        write_failed_exam_list(df_fail, project_name, failed_path)
        print(f"  共 {len(df_fail)} 人考核未通过")
    else:
        print("  无考核未通过人员")
        failed_path = None

    print("✓ 全部完成")
    print(f"\n未完成人员分布:")
    for xq in XQ_ORDER:
        n = len(df[(df['小区负责人']==xq) & (~df['_完成'])])
        print(f"  {xq}: {n} 人")

    return single_path, region_path, east_path, reminder_path, failed_path


if __name__ == '__main__':
    if len(sys.argv) < 4:
        print("用法: python process.py <项目地图情况表> <员工账号管理> <架构表> [输出目录]")
        sys.exit(1)
    out = sys.argv[4] if len(sys.argv) > 4 else '.'
    run(sys.argv[1], sys.argv[2], sys.argv[3], out)
